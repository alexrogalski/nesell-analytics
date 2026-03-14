"""
DPD invoice email ETL: automatically fetch, parse, and import DPD shipping invoices.

DPD Poland sends two types of emails:
1. "E-faktura" from e-faktury@dpd.com.pl -- PDF invoice (aggregated totals, no tracking)
2. "Specyfikacja" from specyfikacje@dpd.com.pl -- XLSX with per-shipment cost breakdown

The XLSX specification files contain:
- Numer Listu (tracking number)
- Data nadania (ship date)
- Cena netto (net cost per line item in PLN)
- Dorecenie (destination country code)
- waga (weight in kg)
- Rodzaj uslugi (service type: base rate, fuel surcharge, security fee, energy fee)
- Nr_Faktury (invoice number)

Each shipment has multiple rows (base + surcharges). This module groups them
by tracking number and sums the total net cost per shipment.

Usage:
    python3.11 -m etl.run --dpd-email          # process new DPD invoice emails
    python3.11 -m etl.run --dpd-email --days 90 # search last 90 days of emails
"""

import imaplib
import email
import io
import re
import tempfile
import time
from collections import defaultdict
from datetime import datetime, timedelta, date
from email.header import decode_header
from pathlib import Path

from . import config, db, fx_rates

# ---------------------------------------------------------------------------
# Gmail IMAP settings
# ---------------------------------------------------------------------------

IMAP_SERVER = "imap.gmail.com"
GMAIL_CREDS = config._load_env_file(config.KEYS_DIR / "nesell-gmail.env")
GMAIL_USER = GMAIL_CREDS.get("GMAIL_USER", "alexander@nesell.co")
GMAIL_PASSWORD = GMAIL_CREDS.get("GMAIL_APP_PASSWORD", "")

# DPD specification emails come from this address
DPD_SPEC_SENDER = "specyfikacje@dpd.com.pl"
# DPD invoice emails come from this address (for cross-reference)
DPD_INVOICE_SENDER = "e-faktury@dpd.com.pl"

# VAT rate for DPD services
DPD_VAT_RATE = 0.23


def _decode_header_value(value: str) -> str:
    """Decode MIME-encoded email header value."""
    if not value:
        return ""
    decoded_parts = decode_header(value)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            result.append(part)
    return "".join(result)


def _connect_imap():
    """Connect to Gmail IMAP and return the mail object."""
    if not GMAIL_PASSWORD:
        raise ValueError(
            "GMAIL_APP_PASSWORD not found in ~/.keys/nesell-gmail.env. "
            "Cannot connect to Gmail IMAP."
        )
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(GMAIL_USER, GMAIL_PASSWORD)
    return mail


def _get_processed_email_uids() -> set[str]:
    """Get set of email UIDs already processed from dpd_invoice_imports table."""
    uids = set()
    offset = 0
    while True:
        rows = db._get("dpd_invoice_imports", {
            "select": "email_uid",
            "limit": "1000",
            "offset": str(offset),
        })
        for r in rows:
            uids.add(str(r["email_uid"]))
        if len(rows) < 1000:
            break
        offset += 1000
    return uids


def _search_dpd_spec_emails(mail, days_back: int = 180) -> list[str]:
    """Search for DPD specification emails (XLSX attachments with per-shipment data).

    Returns list of IMAP message IDs.
    """
    mail.select("INBOX")

    # Search for emails from DPD specyfikacje
    since_date = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    search_criteria = f'(FROM "{DPD_SPEC_SENDER}" SINCE "{since_date}")'

    status, messages = mail.search(None, search_criteria)
    if status != "OK" or not messages[0]:
        return []

    return messages[0].decode().split()


def _extract_xlsx_attachment(msg) -> tuple[bytes | None, str]:
    """Extract XLSX attachment from email message.

    Returns (xlsx_bytes, filename) or (None, "").
    """
    for part in msg.walk():
        filename = part.get_filename()
        if not filename:
            continue

        # Decode filename if MIME-encoded
        filename = _decode_header_value(filename)

        if filename.endswith(".xlsx"):
            content = part.get_payload(decode=True)
            return content, filename

    return None, ""


def _parse_dpd_xlsx(xlsx_bytes: bytes) -> dict[str, dict]:
    """Parse DPD specification XLSX and return per-tracking-number cost data.

    Returns dict: tracking_number -> {
        'cost_net': float (total PLN netto),
        'cost_gross': float (PLN brutto),
        'dest_country': str,
        'weight_kg': float,
        'ship_date': str (YYYY-MM-DD),
        'invoice_number': str,
        'notes': str,
        'service_type': str,
    }
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb.active

    # Read headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else "")

    # Build column index map
    col_idx = {}
    for i, h in enumerate(headers):
        col_idx[h] = i

    # Required columns
    tracking_col = col_idx.get("Numer Listu")
    cost_net_col = col_idx.get("Cena netto")
    service_col = col_idx.get("Rodzaj usługi")
    invoice_col = col_idx.get("Nr_Faktury")
    dest_col = col_idx.get("Doręczenie")
    weight_col = col_idx.get("waga")
    date_col = col_idx.get("Data nadania")
    notes_col = col_idx.get("Uwagi")
    origin_col = col_idx.get("Nadanie")

    if tracking_col is None or cost_net_col is None:
        print(f"    [WARN] Cannot find required columns. Available: {headers}")
        wb.close()
        return {}

    # Parse rows, group costs by tracking number
    tracking_data = defaultdict(lambda: {
        "base_cost": 0.0,
        "fuel_cost": 0.0,
        "security_cost": 0.0,
        "energy_cost": 0.0,
        "other_cost": 0.0,
        "dest_country": "",
        "origin_country": "",
        "weight_kg": 0.0,
        "ship_date": "",
        "invoice_number": "",
        "notes": "",
        "service_type": "",
    })

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(tracking_col, cost_net_col):
            continue

        tracking = str(row[tracking_col] or "").strip()
        if not tracking:
            continue

        cost_net = float(row[cost_net_col] or 0)
        service = str(row[service_col] or "") if service_col is not None else ""
        entry = tracking_data[tracking]

        # Set invoice number
        if invoice_col is not None:
            inv = str(row[invoice_col] or "").strip()
            if inv:
                entry["invoice_number"] = inv

        # Classify the cost line by service type
        service_lower = service.lower()
        if "spedycja" in service_lower or "kurierska" in service_lower:
            entry["base_cost"] += cost_net
            entry["service_type"] = service

            # Capture metadata from the base shipment row
            if dest_col is not None:
                dest = str(row[dest_col] or "").strip()
                if dest:
                    entry["dest_country"] = dest
            if origin_col is not None:
                orig = str(row[origin_col] or "").strip()
                if orig:
                    entry["origin_country"] = orig
            if weight_col is not None:
                wt = float(row[weight_col] or 0)
                if wt > 0:
                    entry["weight_kg"] = wt
            if date_col is not None:
                dt = row[date_col]
                if hasattr(dt, "strftime"):
                    entry["ship_date"] = dt.strftime("%Y-%m-%d")
                elif dt:
                    entry["ship_date"] = str(dt)[:10]
            if notes_col is not None:
                n = str(row[notes_col] or "").strip()
                if n and n != "None":
                    entry["notes"] = n

        elif "paliwow" in service_lower:
            entry["fuel_cost"] += cost_net
        elif "bezpiecze" in service_lower:
            entry["security_cost"] += cost_net
        elif "energetyczn" in service_lower:
            entry["energy_cost"] += cost_net
        elif "potwierdze" in service_lower:
            # POD (proof of delivery) confirmation fee
            entry["other_cost"] += cost_net
        else:
            entry["other_cost"] += cost_net

    wb.close()

    # Aggregate into final per-tracking cost records
    result = {}
    for tracking, data in tracking_data.items():
        total_net = round(
            data["base_cost"] + data["fuel_cost"] + data["security_cost"]
            + data["energy_cost"] + data["other_cost"],
            2,
        )

        # Skip zero-cost lines (can happen with some DPD entries)
        if total_net <= 0:
            continue

        total_gross = round(total_net * (1 + DPD_VAT_RATE), 2)

        result[tracking] = {
            "cost_net": total_net,
            "cost_gross": total_gross,
            "dest_country": data["dest_country"],
            "weight_kg": data["weight_kg"],
            "ship_date": data["ship_date"],
            "invoice_number": data["invoice_number"],
            "notes": data["notes"],
            "service_type": data["service_type"],
            "cost_breakdown": {
                "base": data["base_cost"],
                "fuel": data["fuel_cost"],
                "security": data["security_cost"],
                "energy": data["energy_cost"],
                "other": data["other_cost"],
            },
        }

    return result


def _match_and_update_shipping_costs(
    conn,
    tracking_costs: dict[str, dict],
    invoice_number: str,
) -> tuple[int, int, int]:
    """Match parsed tracking costs to shipping_costs table and update.

    Returns (matched, updated, not_found).
    """
    matched = 0
    updated = 0
    not_found = 0

    for tracking, cost_data in tracking_costs.items():
        # Look up by tracking number in shipping_costs
        rows = db._get("shipping_costs", {
            "select": "id,order_id,external_order_id,cost_source,cost_pln",
            "tracking_number": f"eq.{tracking}",
            "limit": "1",
        })

        if not rows:
            not_found += 1
            continue

        matched += 1
        record = rows[0]

        # Only update if current source is 'estimate' (don't overwrite manual or other invoice data)
        if record.get("cost_source") in ("invoice_email", "invoice_csv", "manual"):
            continue

        # The DPD spec costs are already in PLN netto
        cost_pln = cost_data["cost_gross"]  # We store gross (with VAT) as the total cost

        try:
            update_data = {
                "cost_net": cost_data["cost_net"],
                "cost_gross": cost_data["cost_gross"],
                "cost_currency": "PLN",
                "cost_pln": cost_pln,
                "cost_source": "invoice_email",
                "dpd_invoice_number": invoice_number,
                "updated_at": datetime.now().isoformat(),
            }

            # Update weight if we have it from the spec
            if cost_data.get("weight_kg", 0) > 0:
                update_data["weight_kg"] = cost_data["weight_kg"]

            # Update destination country if available and not already set
            if cost_data.get("dest_country"):
                update_data["destination_country"] = cost_data["dest_country"]

            # Update ship_date if available
            if cost_data.get("ship_date"):
                update_data["ship_date"] = cost_data["ship_date"]

            db._patch("shipping_costs", {"id": f"eq.{record['id']}"}, update_data)

            # Also update the denormalized fields on the orders table
            if record.get("order_id"):
                try:
                    db._patch("orders", {"id": f"eq.{record['order_id']}"}, {
                        "seller_shipping_cost": cost_data["cost_gross"],
                        "seller_shipping_cost_pln": cost_pln,
                    })
                except Exception:
                    pass  # Non-critical

            updated += 1

        except Exception as e:
            print(f"    [WARN] Failed to update tracking {tracking}: {e}")

    return matched, updated, not_found


def _extract_invoice_number_from_subject(subject: str) -> str:
    """Extract DPD invoice number from email subject.

    Subject format: "Specyfikacja DPD 2026-02-28 Platnik: 439631"
    Related invoice: "E-faktura/E-invoice BC10918587 2026-02-28."
    """
    # The XLSX filename contains the invoice number, e.g., BC10918587_2026-02-28_...
    # The subject of spec emails doesn't contain the invoice number directly
    return ""


def _extract_invoice_number_from_filename(filename: str) -> str:
    """Extract DPD invoice number from XLSX filename.

    Format: BC10918587_2026-02-28_20260303204823.xlsx
    """
    match = re.match(r"(BC\d+)", filename)
    return match.group(1) if match else ""


def _extract_period_end_from_subject(subject: str) -> str:
    """Extract period end date from Specyfikacja email subject.

    Format: "Specyfikacja DPD 2026-02-28 Platnik: 439631"
    """
    match = re.search(r"(\d{4}-\d{2}-\d{2})", subject)
    return match.group(1) if match else ""


def sync_dpd_invoices(conn, days_back: int = 180):
    """Main entry point: fetch and process DPD invoice emails from Gmail.

    1. Connect to Gmail IMAP
    2. Search for DPD Specyfikacja emails (XLSX with per-shipment costs)
    3. Filter out already-processed emails
    4. Download and parse XLSX attachments
    5. Match tracking numbers to shipping_costs records
    6. Update costs from estimate to invoice_email
    7. Record processed emails in dpd_invoice_imports

    Args:
        conn: Supabase connection (compatibility, not used directly)
        days_back: How far back to search for emails (default 180 days)

    Returns:
        Total number of shipping cost records updated
    """
    print("  Connecting to Gmail IMAP...")
    try:
        mail = _connect_imap()
    except Exception as e:
        print(f"  [ERROR] Cannot connect to Gmail: {e}")
        return 0

    try:
        # Find DPD specification emails
        print(f"  Searching for DPD Specyfikacja emails (last {days_back} days)...")
        msg_ids = _search_dpd_spec_emails(mail, days_back)
        print(f"  Found {len(msg_ids)} DPD specification email(s)")

        if not msg_ids:
            mail.logout()
            return 0

        # Get already-processed email UIDs
        processed_uids = _get_processed_email_uids()
        print(f"  Already processed: {len(processed_uids)} email(s)")

        total_updated = 0
        total_matched = 0
        total_shipments = 0
        emails_processed = 0

        for msg_id in msg_ids:
            # Fetch the email
            status, data = mail.fetch(msg_id, "(UID RFC822)")
            if status != "OK" or not data or not data[0]:
                continue

            # Extract UID
            uid_match = re.search(rb"UID (\d+)", data[0][0])
            email_uid = uid_match.group(1).decode() if uid_match else msg_id

            # Skip if already processed
            if email_uid in processed_uids:
                continue

            msg = email.message_from_bytes(data[0][1])
            subject = _decode_header_value(msg.get("Subject", ""))
            email_date = msg.get("Date", "")

            print(f"\n  Processing: {subject}")

            # Extract XLSX attachment
            xlsx_bytes, xlsx_filename = _extract_xlsx_attachment(msg)
            if not xlsx_bytes:
                print(f"    [SKIP] No XLSX attachment found")
                # Record as processed anyway (to avoid re-checking)
                _record_import(
                    email_uid, subject, email_date, "", "", xlsx_filename,
                    0, 0, 0, 0, 0, "No XLSX attachment found"
                )
                continue

            print(f"    Attachment: {xlsx_filename} ({len(xlsx_bytes)} bytes)")

            # Extract metadata
            invoice_number = _extract_invoice_number_from_filename(xlsx_filename)
            period_end = _extract_period_end_from_subject(subject)
            print(f"    Invoice: {invoice_number}, Period end: {period_end}")

            # Parse XLSX
            try:
                tracking_costs = _parse_dpd_xlsx(xlsx_bytes)
            except Exception as e:
                print(f"    [ERROR] Failed to parse XLSX: {e}")
                _record_import(
                    email_uid, subject, email_date, invoice_number,
                    period_end, xlsx_filename, 0, 0, 0, 0, 0,
                    f"XLSX parse error: {e}"
                )
                continue

            num_shipments = len(tracking_costs)
            total_net = sum(c["cost_net"] for c in tracking_costs.values())
            total_gross = sum(c["cost_gross"] for c in tracking_costs.values())

            print(f"    Parsed {num_shipments} shipments, total net: {total_net:.2f} PLN")

            # Match and update shipping_costs
            matched, updated, not_found = _match_and_update_shipping_costs(
                conn, tracking_costs, invoice_number
            )

            print(f"    Matched: {matched}, Updated: {updated}, Not in DB: {not_found}")

            # Record the import
            _record_import(
                email_uid, subject, email_date, invoice_number,
                period_end, xlsx_filename, num_shipments, matched, updated,
                total_net, total_gross, None
            )

            total_updated += updated
            total_matched += matched
            total_shipments += num_shipments
            emails_processed += 1

            time.sleep(0.5)  # Be gentle with IMAP

        mail.logout()

        # Summary
        print(f"\n  DPD invoice email sync summary:")
        print(f"    Emails processed:    {emails_processed}")
        print(f"    Total shipments:     {total_shipments}")
        print(f"    Matched to orders:   {total_matched}")
        print(f"    Costs updated:       {total_updated}")

        return total_updated

    except Exception as e:
        print(f"  [ERROR] DPD invoice sync failed: {e}")
        import traceback
        traceback.print_exc()
        try:
            mail.logout()
        except Exception:
            pass
        raise


def _record_import(
    email_uid: str,
    subject: str,
    email_date: str,
    invoice_number: str,
    period_end: str,
    filename: str,
    shipments_found: int,
    shipments_matched: int,
    shipments_updated: int,
    total_net: float,
    total_gross: float,
    error_message: str | None,
):
    """Record a processed email in dpd_invoice_imports table."""
    record = {
        "email_uid": str(email_uid),
        "email_subject": subject[:500] if subject else "",
        "invoice_number": invoice_number or None,
        "attachment_filename": filename or None,
        "shipments_found": shipments_found,
        "shipments_matched": shipments_matched,
        "shipments_updated": shipments_updated,
        "total_net_pln": round(total_net, 2),
        "total_gross_pln": round(total_gross, 2),
        "error_message": error_message,
    }

    # Parse email_date to ISO format
    if email_date:
        try:
            from email.utils import parsedate_to_datetime
            dt = parsedate_to_datetime(email_date)
            record["email_date"] = dt.isoformat()
        except Exception:
            pass

    # Parse period_end to date
    if period_end:
        record["invoice_period_end"] = period_end

    try:
        db._post("dpd_invoice_imports", [record], on_conflict="email_uid")
    except Exception as e:
        print(f"    [WARN] Failed to record import: {e}")
