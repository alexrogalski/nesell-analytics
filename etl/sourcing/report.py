"""Generate sourcing analysis reports in Excel, CSV, and terminal formats.

Three output modes:

* **Excel** (``generate_excel_report``) -- multi-sheet workbook with
  summary, profitable products, per-market breakdown, and rejected items.
* **CSV** (``generate_csv``) -- flat file for easy import into other tools.
* **Terminal** (``print_terminal_summary``) -- coloured table for the CLI.

Usage::

    from etl.sourcing.report import (
        generate_excel_report,
        generate_csv,
        print_terminal_summary,
    )

    generate_excel_report(results, recommendations, path="report.xlsx")
    generate_csv(results, path="report.csv")
    print_terminal_summary(results, recommendations)
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import Sequence

from .profit_calculator import ProfitAnalysis
from .quantity_recommender import PurchaseRecommendation

# ---------------------------------------------------------------------------
# Colour helpers (ANSI)
# ---------------------------------------------------------------------------

_GREEN = "\033[92m"
_YELLOW = "\033[93m"
_RED = "\033[91m"
_BOLD = "\033[1m"
_DIM = "\033[2m"
_RESET = "\033[0m"


def _verdict_colour(verdict: str) -> str:
    if verdict == "EXCELLENT":
        return _GREEN
    if verdict == "GOOD":
        return _GREEN
    if verdict == "MARGINAL":
        return _YELLOW
    return _RED


def _risk_colour(risk: str) -> str:
    if risk == "LOW":
        return _GREEN
    if risk == "MEDIUM":
        return _YELLOW
    return _RED


# ---------------------------------------------------------------------------
# Terminal summary
# ---------------------------------------------------------------------------

def print_terminal_summary(
    results: dict[str, list[ProfitAnalysis]],
    recommendations: dict[str, PurchaseRecommendation] | None = None,
    min_margin: float = 0.0,
    file=None,
) -> None:
    """Print a human-readable summary to the terminal.

    Parameters
    ----------
    results : dict[str, list[ProfitAnalysis]]
        ``{ean: [ProfitAnalysis, ...]}`` mapping.
    recommendations : dict[str, PurchaseRecommendation] | None
        ``{ean: PurchaseRecommendation}`` mapping (optional).
    min_margin : float
        Only display products where at least one market exceeds this margin.
    file
        Output stream (defaults to ``sys.stdout``).
    """
    out = file or sys.stdout

    total_products = len(results)
    profitable_eans = set()
    all_analyses: list[ProfitAnalysis] = []

    for ean, analyses in results.items():
        for a in analyses:
            all_analyses.append(a)
            if a.margin_pct >= max(min_margin, 10.0) and not a.errors:
                profitable_eans.add(ean)

    total_profitable = len(profitable_eans)
    total_markets = len(all_analyses)

    # Header
    out.write(f"\n{_BOLD}{'=' * 72}{_RESET}\n")
    out.write(f"{_BOLD}  SOURCING ANALYSIS REPORT  {_DIM}{datetime.now().strftime('%Y-%m-%d %H:%M')}{_RESET}\n")
    out.write(f"{_BOLD}{'=' * 72}{_RESET}\n\n")

    out.write(f"  Products analysed:   {total_products}\n")
    out.write(f"  Market combinations: {total_markets}\n")
    out.write(
        f"  Profitable (>={max(min_margin, 10.0):.0f}%):  "
        f"{_GREEN}{total_profitable}{_RESET} / {total_products}\n"
    )
    out.write("\n")

    # Profitable products table
    if profitable_eans:
        out.write(f"{_BOLD}  TOP PROFITABLE PRODUCTS{_RESET}\n")
        out.write(f"  {'EAN':<15} {'Best Market':<14} {'Sell':>8} {'Cost':>8} "
                  f"{'Profit':>8} {'Margin':>7} {'ROI':>7} {'Verdict':<12}\n")
        out.write(f"  {'-' * 15} {'-' * 14} {'-' * 8} {'-' * 8} "
                  f"{'-' * 8} {'-' * 7} {'-' * 7} {'-' * 12}\n")

        # Sort profitable by best margin.
        sorted_eans = sorted(
            profitable_eans,
            key=lambda e: max(
                (a.profit_pln for a in results[e] if not a.errors),
                default=0,
            ),
            reverse=True,
        )

        for ean in sorted_eans[:30]:
            best = max(
                (a for a in results[ean] if not a.errors),
                key=lambda a: a.profit_pln,
                default=None,
            )
            if best is None:
                continue

            clr = _verdict_colour(best.verdict)
            out.write(
                f"  {ean:<15} {best.platform:<14} "
                f"{best.sell_price_pln:>8.2f} "
                f"{best.total_costs_pln:>8.2f} "
                f"{clr}{best.profit_pln:>8.2f}{_RESET} "
                f"{clr}{best.margin_pct:>6.1f}%{_RESET} "
                f"{best.roi_pct:>6.1f}% "
                f"{clr}{best.verdict:<12}{_RESET}\n"
            )

        out.write("\n")

    # Recommendations summary
    if recommendations:
        recs = [r for r in recommendations.values() if r.recommended_qty > 0]
        if recs:
            recs.sort(key=lambda r: r.estimated_monthly_profit_pln, reverse=True)

            total_invest = sum(r.total_investment_pln for r in recs)
            total_monthly_profit = sum(r.estimated_monthly_profit_pln for r in recs)

            out.write(f"{_BOLD}  PURCHASE RECOMMENDATIONS{_RESET}\n")
            out.write(f"  Total investment: {total_invest:,.2f} PLN\n")
            out.write(f"  Est. monthly profit: {total_monthly_profit:,.2f} PLN\n\n")

            out.write(f"  {'EAN':<15} {'Name':<25} {'Qty':>5} {'Invest':>10} "
                      f"{'Profit/mo':>10} {'Risk':<8}\n")
            out.write(f"  {'-' * 15} {'-' * 25} {'-' * 5} {'-' * 10} "
                      f"{'-' * 10} {'-' * 8}\n")

            for r in recs[:30]:
                rclr = _risk_colour(r.risk_score)
                name_trunc = (r.name[:23] + "..") if len(r.name) > 25 else r.name
                out.write(
                    f"  {r.ean:<15} {name_trunc:<25} {r.recommended_qty:>5} "
                    f"{r.total_investment_pln:>10,.2f} "
                    f"{r.estimated_monthly_profit_pln:>10,.2f} "
                    f"{rclr}{r.risk_score:<8}{_RESET}\n"
                )

            out.write("\n")

    out.write(f"{_DIM}  Report generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}{_RESET}\n\n")


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def generate_excel_report(
    results: dict[str, list[ProfitAnalysis]],
    recommendations: dict[str, PurchaseRecommendation] | None = None,
    path: str = "sourcing_report.xlsx",
    min_margin: float = 10.0,
) -> str:
    """Write a multi-sheet Excel workbook and return the file path.

    Sheets:
        1. Summary -- aggregate stats.
        2. Profitable -- products above *min_margin*, sorted by ROI.
        3. All Markets -- per-EAN, per-market full breakdown.
        4. Rejected -- products below the margin threshold.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel report generation. "
            "Install with: pip install openpyxl"
        )

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    pct_format = '0.0"%"'
    money_format = '#,##0.00'

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # Flatten all analyses.
    all_analyses: list[ProfitAnalysis] = []
    for ean_analyses in results.values():
        all_analyses.extend(ean_analyses)

    profitable_all = [a for a in all_analyses if a.margin_pct >= min_margin and not a.errors]
    rejected_all = [a for a in all_analyses if a.margin_pct < min_margin or a.errors]

    # Unique profitable EANs.
    profitable_eans = set()
    for a in profitable_all:
        profitable_eans.add(a.ean)

    # -------------------------------------------------------------------
    # Sheet 1: Summary
    # -------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary"

    summary_data = [
        ("Sourcing Analysis Report", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Products Analysed", len(results)),
        ("Profitable Products", len(profitable_eans)),
        ("Total Market Combinations", len(all_analyses)),
        ("Profitable Combinations", len(profitable_all)),
        ("Rejected Combinations", len(rejected_all)),
    ]

    if recommendations:
        recs = [r for r in recommendations.values() if r.recommended_qty > 0]
        total_invest = sum(r.total_investment_pln for r in recs)
        total_profit = sum(r.estimated_monthly_profit_pln for r in recs)
        summary_data.extend([
            ("", ""),
            ("Recommended Products to Buy", len(recs)),
            ("Total Investment (PLN)", round(total_invest, 2)),
            ("Est. Monthly Profit (PLN)", round(total_profit, 2)),
        ])

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if row_idx <= 2 else Font()
        cell = ws1.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = money_format

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20

    # -------------------------------------------------------------------
    # Sheet 2: Profitable
    # -------------------------------------------------------------------
    ws2 = wb.create_sheet("Profitable")
    headers2 = [
        "EAN", "ASIN", "Name", "Best Market", "Sell Price", "Currency",
        "Sell PLN", "Purchase PLN", "Fees PLN", "Shipping PLN",
        "Profit PLN", "Margin %", "ROI %", "BSR", "Competition",
        "Monthly Sales", "Verdict",
    ]
    _write_header(ws2, headers2)

    # One row per profitable EAN, showing the best market.
    best_per_ean: dict[str, ProfitAnalysis] = {}
    for a in profitable_all:
        if a.ean not in best_per_ean or a.profit_pln > best_per_ean[a.ean].profit_pln:
            best_per_ean[a.ean] = a

    sorted_profitable = sorted(best_per_ean.values(), key=lambda a: a.roi_pct, reverse=True)

    for row_idx, a in enumerate(sorted_profitable, 2):
        ws2.cell(row=row_idx, column=1, value=a.ean)
        ws2.cell(row=row_idx, column=2, value=a.asin)
        ws2.cell(row=row_idx, column=3, value=a.title[:60] if a.title else "")
        ws2.cell(row=row_idx, column=4, value=a.platform)
        ws2.cell(row=row_idx, column=5, value=a.sell_price).number_format = money_format
        ws2.cell(row=row_idx, column=6, value=a.sell_currency)
        ws2.cell(row=row_idx, column=7, value=a.sell_price_pln).number_format = money_format
        ws2.cell(row=row_idx, column=8, value=a.purchase_price_pln).number_format = money_format
        ws2.cell(row=row_idx, column=9, value=a.platform_fees_pln).number_format = money_format
        ws2.cell(row=row_idx, column=10, value=a.shipping_cost_pln).number_format = money_format
        ws2.cell(row=row_idx, column=11, value=a.profit_pln).number_format = money_format
        ws2.cell(row=row_idx, column=12, value=a.margin_pct).number_format = pct_format
        ws2.cell(row=row_idx, column=13, value=a.roi_pct).number_format = pct_format
        ws2.cell(row=row_idx, column=14, value=a.bsr_rank if a.bsr_rank else "")
        ws2.cell(row=row_idx, column=15, value=a.competition)
        ws2.cell(row=row_idx, column=16, value=a.estimated_monthly_sales)
        ws2.cell(row=row_idx, column=17, value=a.verdict)

    # Auto-width approximation.
    for col_idx, _ in enumerate(headers2, 1):
        ws2.column_dimensions[chr(64 + min(col_idx, 26))].width = 14

    # -------------------------------------------------------------------
    # Sheet 3: All Markets
    # -------------------------------------------------------------------
    ws3 = wb.create_sheet("All Markets")
    headers3 = [
        "EAN", "Platform", "ASIN", "Title", "Sell Price", "Currency",
        "Sell PLN", "Purchase PLN", "Fees PLN", "Shipping PLN",
        "Profit PLN", "Margin %", "ROI %", "BSR", "Competition", "Verdict",
        "Errors",
    ]
    _write_header(ws3, headers3)

    # Sort by EAN then by profit desc.
    sorted_all = sorted(all_analyses, key=lambda a: (a.ean, -a.profit_pln))

    for row_idx, a in enumerate(sorted_all, 2):
        ws3.cell(row=row_idx, column=1, value=a.ean)
        ws3.cell(row=row_idx, column=2, value=a.platform)
        ws3.cell(row=row_idx, column=3, value=a.asin)
        ws3.cell(row=row_idx, column=4, value=a.title[:50] if a.title else "")
        ws3.cell(row=row_idx, column=5, value=a.sell_price).number_format = money_format
        ws3.cell(row=row_idx, column=6, value=a.sell_currency)
        ws3.cell(row=row_idx, column=7, value=a.sell_price_pln).number_format = money_format
        ws3.cell(row=row_idx, column=8, value=a.purchase_price_pln).number_format = money_format
        ws3.cell(row=row_idx, column=9, value=a.platform_fees_pln).number_format = money_format
        ws3.cell(row=row_idx, column=10, value=a.shipping_cost_pln).number_format = money_format
        ws3.cell(row=row_idx, column=11, value=a.profit_pln).number_format = money_format
        ws3.cell(row=row_idx, column=12, value=a.margin_pct).number_format = pct_format
        ws3.cell(row=row_idx, column=13, value=a.roi_pct).number_format = pct_format
        ws3.cell(row=row_idx, column=14, value=a.bsr_rank if a.bsr_rank else "")
        ws3.cell(row=row_idx, column=15, value=a.competition)
        ws3.cell(row=row_idx, column=16, value=a.verdict)
        ws3.cell(row=row_idx, column=17, value="; ".join(a.errors) if a.errors else "")

    for col_idx, _ in enumerate(headers3, 1):
        ws3.column_dimensions[chr(64 + min(col_idx, 26))].width = 14

    # -------------------------------------------------------------------
    # Sheet 4: Rejected
    # -------------------------------------------------------------------
    ws4 = wb.create_sheet("Rejected")
    headers4 = [
        "EAN", "Platform", "Sell PLN", "Purchase PLN", "Profit PLN",
        "Margin %", "Verdict", "Errors",
    ]
    _write_header(ws4, headers4)

    sorted_rejected = sorted(rejected_all, key=lambda a: a.margin_pct, reverse=True)

    for row_idx, a in enumerate(sorted_rejected, 2):
        ws4.cell(row=row_idx, column=1, value=a.ean)
        ws4.cell(row=row_idx, column=2, value=a.platform)
        ws4.cell(row=row_idx, column=3, value=a.sell_price_pln).number_format = money_format
        ws4.cell(row=row_idx, column=4, value=a.purchase_price_pln).number_format = money_format
        ws4.cell(row=row_idx, column=5, value=a.profit_pln).number_format = money_format
        ws4.cell(row=row_idx, column=6, value=a.margin_pct).number_format = pct_format
        ws4.cell(row=row_idx, column=7, value=a.verdict)
        ws4.cell(row=row_idx, column=8, value="; ".join(a.errors) if a.errors else "")

    # Save
    out_path = Path(path).resolve()
    wb.save(str(out_path))
    return str(out_path)


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

_CSV_HEADERS = [
    "ean", "platform", "asin", "title", "sell_price", "sell_currency",
    "sell_price_pln", "purchase_price_pln", "platform_fees_pln",
    "shipping_cost_pln", "profit_pln", "margin_pct", "roi_pct",
    "bsr_rank", "competition", "estimated_monthly_sales", "verdict",
]


def generate_csv(
    results: dict[str, list[ProfitAnalysis]],
    path: str = "sourcing_report.csv",
) -> str:
    """Write a flat CSV with one row per EAN/market combination.

    Returns the resolved file path.
    """
    out_path = Path(path).resolve()

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=_CSV_HEADERS, extrasaction="ignore")
        writer.writeheader()

        for ean, analyses in sorted(results.items()):
            for a in sorted(analyses, key=lambda x: -x.profit_pln):
                writer.writerow({
                    "ean": a.ean,
                    "platform": a.platform,
                    "asin": a.asin,
                    "title": a.title,
                    "sell_price": a.sell_price,
                    "sell_currency": a.sell_currency,
                    "sell_price_pln": a.sell_price_pln,
                    "purchase_price_pln": a.purchase_price_pln,
                    "platform_fees_pln": a.platform_fees_pln,
                    "shipping_cost_pln": a.shipping_cost_pln,
                    "profit_pln": a.profit_pln,
                    "margin_pct": a.margin_pct,
                    "roi_pct": a.roi_pct,
                    "bsr_rank": a.bsr_rank or "",
                    "competition": a.competition,
                    "estimated_monthly_sales": a.estimated_monthly_sales,
                    "verdict": a.verdict,
                })

    return str(out_path)
