"""Parse supplier CSV/XLSX price lists into normalised SupplierProduct records.

Handles Polish and English column names, validates EAN-13 check digits,
and supports both .csv and .xlsx inputs via openpyxl.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class SupplierProduct:
    ean: str
    purchase_price: float
    purchase_currency: str = "PLN"
    name: str = ""
    supplier_sku: str = ""
    weight_kg: float | None = None
    moq: int = 1
    available_qty: int | None = None


# ---------------------------------------------------------------------------
# Column-name aliases (lowercase, stripped)
# ---------------------------------------------------------------------------

_ALIASES: dict[str, list[str]] = {
    "ean": [
        "ean", "gtin", "barcode", "kod_kreskowy", "kod_ean", "ean13",
        "ean-13", "ean_13", "kod ean", "kod kreskowy",
    ],
    "price": [
        "cena", "price", "cena_netto", "cena_zakupu", "purchase_price",
        "cost", "koszt", "cena netto", "cena zakupu",
    ],
    "currency": [
        "waluta", "currency",
    ],
    "name": [
        "nazwa", "name", "produkt", "product", "opis", "description",
        "nazwa produktu", "product name",
    ],
    "weight": [
        "waga", "weight", "masa", "waga_kg", "weight_kg",
    ],
    "sku": [
        "sku", "supplier_sku", "indeks", "kod_produktu", "kod produktu",
        "numer katalogowy",
    ],
    "moq": [
        "moq", "min_order", "minimalne_zamowienie", "min order",
    ],
    "available_qty": [
        "dostepnosc", "stan", "qty", "available", "ilosc", "available_qty",
        "stock",
    ],
}


def _build_lookup() -> dict[str, str]:
    """Return {alias: canonical_field} mapping."""
    lookup: dict[str, str] = {}
    for canonical, aliases in _ALIASES.items():
        for alias in aliases:
            lookup[alias] = canonical
    return lookup


_ALIAS_LOOKUP = _build_lookup()


# ---------------------------------------------------------------------------
# EAN-13 validation
# ---------------------------------------------------------------------------

def validate_ean13(code: str) -> bool:
    """Return True if *code* is a valid EAN-13 (correct length and check digit)."""
    if not code or not code.isdigit() or len(code) != 13:
        return False
    digits = [int(d) for d in code]
    total = sum(d * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits[:12]))
    check = (10 - total % 10) % 10
    return check == digits[12]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalise_header(raw: str) -> str:
    """Lowercase, strip whitespace, collapse separators."""
    s = raw.strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    return s


def _resolve_columns(headers: Sequence[str]) -> dict[str, int]:
    """Map canonical field names to column indices using the alias table."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        norm = _normalise_header(raw)
        # Try the normalised header directly, then without underscores
        canonical = _ALIAS_LOOKUP.get(norm)
        if canonical is None:
            canonical = _ALIAS_LOOKUP.get(norm.replace("_", " "))
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


def _safe_float(value: object) -> float | None:
    """Convert a cell value to float, tolerating commas and whitespace."""
    if value is None:
        return None
    s = str(value).strip().replace(",", ".").replace(" ", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _safe_int(value: object) -> int | None:
    if value is None:
        return None
    f = _safe_float(value)
    if f is None:
        return None
    return int(f)


def _clean_ean(value: object) -> str:
    """Extract a 13-digit string from a cell (handles floats like 5.90126E+12)."""
    if value is None:
        return ""
    s = str(value).strip()
    # Excel sometimes stores EANs as floats: 5901234567890.0
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    # Strip non-digit characters
    digits = re.sub(r"\D", "", s)
    # Pad to 13 if 12 digits (missing leading zero)
    if len(digits) == 12:
        digits = "0" + digits
    return digits


# ---------------------------------------------------------------------------
# Row-to-product conversion
# ---------------------------------------------------------------------------

def _row_to_product(
    row: Sequence[object],
    col_map: dict[str, int],
    default_currency: str,
) -> SupplierProduct | None:
    """Convert a single row into a SupplierProduct. Returns None on failure."""
    ean_idx = col_map.get("ean")
    price_idx = col_map.get("price")

    if ean_idx is None or price_idx is None:
        return None

    ean = _clean_ean(row[ean_idx]) if ean_idx < len(row) else ""
    if not validate_ean13(ean):
        return None

    price = _safe_float(row[price_idx] if price_idx < len(row) else None)
    if price is None or price <= 0:
        return None

    currency = default_currency
    if "currency" in col_map:
        cidx = col_map["currency"]
        if cidx < len(row) and row[cidx]:
            currency = str(row[cidx]).strip().upper() or default_currency

    name = ""
    if "name" in col_map:
        nidx = col_map["name"]
        if nidx < len(row) and row[nidx]:
            name = str(row[nidx]).strip()

    sku = ""
    if "sku" in col_map:
        sidx = col_map["sku"]
        if sidx < len(row) and row[sidx]:
            sku = str(row[sidx]).strip()

    weight = None
    if "weight" in col_map:
        widx = col_map["weight"]
        weight = _safe_float(row[widx] if widx < len(row) else None)

    moq = 1
    if "moq" in col_map:
        midx = col_map["moq"]
        moq_val = _safe_int(row[midx] if midx < len(row) else None)
        if moq_val and moq_val > 0:
            moq = moq_val

    available_qty = None
    if "available_qty" in col_map:
        aidx = col_map["available_qty"]
        available_qty = _safe_int(row[aidx] if aidx < len(row) else None)

    return SupplierProduct(
        ean=ean,
        purchase_price=price,
        purchase_currency=currency,
        name=name,
        supplier_sku=sku,
        weight_kg=weight,
        moq=moq,
        available_qty=available_qty,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_file(
    path: str | Path,
    default_currency: str = "PLN",
    sheet_name: str | None = None,
) -> list[SupplierProduct]:
    """Parse a supplier price list (CSV or XLSX) and return valid products.

    Parameters
    ----------
    path : str | Path
        Path to the .csv or .xlsx file.
    default_currency : str
        Fallback currency when the file has no currency column.
    sheet_name : str | None
        For XLSX files, the worksheet to read. Defaults to the first sheet.

    Returns
    -------
    list[SupplierProduct]
        Only rows with valid EAN-13 and positive price are included.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return _parse_xlsx(path, default_currency, sheet_name)
    elif suffix in (".csv", ".tsv", ".txt"):
        return _parse_csv(path, default_currency)
    else:
        raise ValueError(f"Unsupported file format: {suffix}. Use .csv or .xlsx.")


def _parse_csv(path: Path, default_currency: str) -> list[SupplierProduct]:
    """Parse a CSV/TSV file."""
    raw = path.read_bytes()

    # Detect encoding (try utf-8 first, fall back to cp1250 for Polish files)
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("latin-1")

    # Detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)

    if len(rows) < 2:
        return []

    col_map = _resolve_columns(rows[0])
    if "ean" not in col_map or "price" not in col_map:
        raise ValueError(
            f"Cannot find required columns (EAN + price). "
            f"Detected headers: {rows[0]}"
        )

    products: list[SupplierProduct] = []
    for row in rows[1:]:
        p = _row_to_product(row, col_map, default_currency)
        if p is not None:
            products.append(p)

    return products


def _parse_xlsx(
    path: Path,
    default_currency: str,
    sheet_name: str | None,
) -> list[SupplierProduct]:
    """Parse an XLSX file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX parsing. Install with: "
            "pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return []

    headers = [str(c) if c is not None else "" for c in header_row]
    col_map = _resolve_columns(headers)

    if "ean" not in col_map or "price" not in col_map:
        wb.close()
        raise ValueError(
            f"Cannot find required columns (EAN + price). "
            f"Detected headers: {headers}"
        )

    products: list[SupplierProduct] = []
    for row in rows_iter:
        cells = list(row)
        p = _row_to_product(cells, col_map, default_currency)
        if p is not None:
            products.append(p)

    wb.close()
    return products
